from openpyxl import load_workbook

_WB = None
_WS_DATA = None

def _load_excel_once(excel_path: str):
    global _WB, _WS_DATA

    if _WB is None:
        _WB = load_workbook(excel_path, data_only=True)
        _WS_DATA = _WB["Data"]


def get_best_result_per_unit_from_excel(
    excel_path: str,
    airUnit: str,
    airflow: float,
    spUnit: str,
    pressure: float,
    modelType: str,
    motorType: str,
    sreInput: float,
    moistInput: float,
):
    
    
    # wb = load_workbook(excel_path, data_only=True)
    # ws_data = wb["Data"]

    _load_excel_once(excel_path)
    ws_data = _WS_DATA

    # basic validation (same as VBA)
    if not (modelType and motorType and airUnit and spUnit):
        return {"error": "Please complete all inputs."}
    if modelType == "HRV" and moistInput > 1:
        return {
            "error": "Invalid Input: Net Moisture Transfer cannot exceed 1 for HRV models."
        }

    # tolerances
    TOL_AIR = 2.0
    TOL_MOIST = 7.0
    TOL_SRE = 2.0

    if spUnit == "Pascals (Pa)":
        spTol = 6.25
    else:
        spTol = 0.05

    # find "Product Details and Files" column
    linkCol = None
    for c in range(1, ws_data.max_column + 1):
        if ws_data.cell(row=1, column=c).value == "Product Details and Files":
            linkCol = c
            break
    if linkCol is None:
        return {"error": "Cannot find 'Product Details and Files' header in Data sheet."}

    lastRow = ws_data.max_row

    # collect unique models for given modelType + motorType
    models = {}
    for r in range(2, lastRow + 1):
        if (
            ws_data.cell(row=r, column=2).value == modelType
            and ws_data.cell(row=r, column=3).value == motorType
        ):
            modelName = ws_data.cell(row=r, column=1).value
            if modelName and modelName not in models:
                models[modelName] = True

    if not models:
        return {"error": "No units found for this Model Type and Motor Type."}

    unit_keys = list(models.keys())
    unitCount = len(unit_keys)

    arrRow = [0] * unitCount
    arrBlock = [0] * unitCount
    arrIsBase = [False] * unitCount
    arrSRE = [0.0] * unitCount
    arrAirDiff = [0.0] * unitCount

    idxMatch = 0
    skippedUnits = 0

    for unitKey in unit_keys:
        # strict candidate
        bestRow = 0
        bestBlockCol = 0
        bestIsBase = True
        bestSRE = -1.0
        bestAirDiff = 1e20

        # fallback candidate
        fbRow = 0
        fbBlockCol = 0
        fbIsBase = True
        fbSRE = -1.0
        fbAirDiff = 1e20
        fbBestMoistDelta = 1e20
        fbBestSreDelta = 1e20

        for r in range(2, lastRow + 1):
            if (
                ws_data.cell(row=r, column=1).value == unitKey
                and ws_data.cell(row=r, column=2).value == modelType
                and ws_data.cell(row=r, column=3).value == motorType
            ):

                # ---- BASE BLOCK F–L (cols 6–12) ----
                inwg = float(ws_data.cell(row=r, column=6).value or 0)
                cfm = float(ws_data.cell(row=r, column=7).value or 0)
                pa = float(ws_data.cell(row=r, column=8).value or 0)
                ls = float(ws_data.cell(row=r, column=9).value or 0)
                watts = float(ws_data.cell(row=r, column=10).value or 0)
                sreVal = float(ws_data.cell(row=r, column=11).value or 0)
                moistVal = float(ws_data.cell(row=r, column=12).value or 0)

                if spUnit == "Pascals (Pa)":
                    spCompare = pa
                else:
                    spCompare = inwg

                if airUnit == "Cubic Feet Per Minute (CFM)":
                    airCompare = cfm
                else:
                    airCompare = ls

                # STRICT
                if (
                    abs(spCompare - pressure) <= spTol
                    and abs(airCompare - airflow) <= TOL_AIR
                    and abs(moistVal - moistInput) <= TOL_MOIST
                    and abs(sreVal - sreInput) <= TOL_SRE
                ):
                    airDiff = abs(airCompare - airflow)
                    if (sreVal > bestSRE) or (
                        sreVal == bestSRE and airDiff < bestAirDiff
                    ):
                        bestSRE = sreVal
                        bestAirDiff = airDiff
                        bestRow = r
                        bestBlockCol = 0
                        bestIsBase = True

                # FALLBACK
                if (
                    abs(spCompare - pressure) <= spTol
                    and abs(airCompare - airflow) <= TOL_AIR
                    and moistVal >= moistInput
                    and sreVal >= sreInput
                ):
                    airDiff = abs(airCompare - airflow)
                    moistDelta = moistVal - moistInput
                    sreDelta = sreVal - sreInput
                    if (
                        moistDelta < fbBestMoistDelta
                        or (
                            moistDelta == fbBestMoistDelta
                            and sreDelta < fbBestSreDelta
                        )
                        or (
                            moistDelta == fbBestMoistDelta
                            and sreDelta == fbBestSreDelta
                            and airDiff < fbAirDiff
                        )
                    ):
                        fbBestMoistDelta = moistDelta
                        fbBestSreDelta = sreDelta
                        fbAirDiff = airDiff
                        fbRow = r
                        fbBlockCol = 0
                        fbIsBase = True
                        fbSRE = sreVal

                # ---- REPEATED 7-COLUMN BLOCKS (M–S, T–Z, ... TO BEFORE linkCol) ----
                maxBlockStart = linkCol - 7
                for c in range(13, maxBlockStart + 1, 7):
                    inwg = float(ws_data.cell(row=r, column=c).value or 0)
                    cfm = float(ws_data.cell(row=r, column=c + 1).value or 0)
                    pa = float(ws_data.cell(row=r, column=c + 2).value or 0)
                    ls = float(ws_data.cell(row=r, column=c + 3).value or 0)
                    watts = float(ws_data.cell(row=r, column=c + 4).value or 0)
                    sreVal = float(ws_data.cell(row=r, column=c + 5).value or 0)
                    moistVal = float(ws_data.cell(row=r, column=c + 6).value or 0)

                    if spUnit == "Pascals (Pa)":
                        spCompare = pa
                    else:
                        spCompare = inwg

                    if airUnit == "Cubic Feet Per Minute (CFM)":
                        airCompare = cfm
                    else:
                        airCompare = ls

                    # STRICT
                    if (
                        abs(spCompare - pressure) <= spTol
                        and abs(airCompare - airflow) <= TOL_AIR
                        and abs(moistVal - moistInput) <= TOL_MOIST
                        and abs(sreVal - sreInput) <= TOL_SRE
                    ):
                        airDiff = abs(airCompare - airflow)
                        if (sreVal > bestSRE) or (
                            sreVal == bestSRE and airDiff < bestAirDiff
                        ):
                            bestSRE = sreVal
                            bestAirDiff = airDiff
                            bestRow = r
                            bestBlockCol = c
                            bestIsBase = False

                    # FALLBACK
                    if (
                        abs(spCompare - pressure) <= spTol
                        and abs(airCompare - airflow) <= TOL_AIR
                        and moistVal >= moistInput
                        and sreVal >= sreInput
                    ):
                        airDiff = abs(airCompare - airflow)
                        moistDelta = moistVal - moistInput
                        sreDelta = sreVal - sreInput
                        if (
                            moistDelta < fbBestMoistDelta
                            or (
                                moistDelta == fbBestMoistDelta
                                and sreDelta < fbBestSreDelta
                            )
                            or (
                                moistDelta == fbBestMoistDelta
                                and sreDelta == fbBestSreDelta
                                and airDiff < fbAirDiff
                            )
                        ):
                            fbBestMoistDelta = moistDelta
                            fbBestSreDelta = sreDelta
                            fbAirDiff = airDiff
                            fbRow = r
                            fbBlockCol = c
                            fbIsBase = False
                            fbSRE = sreVal

        # decide final candidate for this unit
        if bestRow > 0:
            finalRow = bestRow
            finalBlock = bestBlockCol
            finalIsBase = bestIsBase
            finalSRE = bestSRE
            finalAirDiff = bestAirDiff
        elif fbRow > 0:
            finalRow = fbRow
            finalBlock = fbBlockCol
            finalIsBase = fbIsBase
            finalSRE = fbSRE
            finalAirDiff = fbAirDiff
        else:
            finalRow = 0

        if finalRow > 0:
            arrRow[idxMatch] = finalRow
            arrBlock[idxMatch] = finalBlock
            arrIsBase[idxMatch] = finalIsBase
            arrSRE[idxMatch] = finalSRE
            arrAirDiff[idxMatch] = finalAirDiff
            idxMatch += 1
        else:
            skippedUnits += 1

    if idxMatch == 0:
        return {"error": "No matching units found within tolerance."}

    # trim arrays
    arrRow = arrRow[:idxMatch]
    arrBlock = arrBlock[:idxMatch]
    arrIsBase = arrIsBase[:idxMatch]
    arrSRE = arrSRE[:idxMatch]
    arrAirDiff = arrAirDiff[:idxMatch]

    # sort: SRE desc, then AirDiff asc
    indices = list(range(idxMatch))
    indices.sort(key=lambda i: (-arrSRE[i], arrAirDiff[i]))

    results = []
    for idx in indices:
        r = arrRow[idx]
        block_col = arrBlock[idx]
        is_base = arrIsBase[idx]

        model = ws_data.cell(row=r, column=1).value
        model_type_val = ws_data.cell(row=r, column=2).value
        motor_type_val = ws_data.cell(row=r, column=3).value
        mca = ws_data.cell(row=r, column=4).value
        mocp = ws_data.cell(row=r, column=5).value

        if is_base or block_col == 0:
            inwg = float(ws_data.cell(row=r, column=6).value or 0)
            cfm = float(ws_data.cell(row=r, column=7).value or 0)
            pa = float(ws_data.cell(row=r, column=8).value or 0)
            ls = float(ws_data.cell(row=r, column=9).value or 0)
            watts = float(ws_data.cell(row=r, column=10).value or 0)
            sreVal = float(ws_data.cell(row=r, column=11).value or 0)
            moistVal = float(ws_data.cell(row=r, column=12).value or 0)
        else:
            c = block_col
            inwg = float(ws_data.cell(row=r, column=c).value or 0)
            cfm = float(ws_data.cell(row=r, column=c + 1).value or 0)
            pa = float(ws_data.cell(row=r, column=c + 2).value or 0)
            ls = float(ws_data.cell(row=r, column=c + 3).value or 0)
            watts = float(ws_data.cell(row=r, column=c + 4).value or 0)
            sreVal = float(ws_data.cell(row=r, column=c + 5).value or 0)
            moistVal = float(ws_data.cell(row=r, column=c + 6).value or 0)

        pa_out = round(pa, 2)
        inwg_out = round(inwg, 3)
        ls_out = round(ls, 3)
        cfm_out = round(cfm, 0)

        link_cell = ws_data.cell(row=r, column=linkCol)
        linkText = link_cell.value
        linkAddress = None
        if link_cell.hyperlink is not None:
            linkAddress = link_cell.hyperlink.target

        results.append(
            {
                "Model": model,
    "Model Type": model_type_val,
    "Motor Type": motor_type_val,
    "MCA": round(float(mca or 0), 2),
    "MOCP": round(float(mocp or 0), 2),

    "Static Pressure (Pa)": round(pa, 2),
    "Static Pressure (IN W.G.)": round(inwg, 3),

    "Net Supply (L/S)": round(ls, 3),
    "Net Supply (CFM)": int(round(cfm, 0)),

    "Watts": int(round(watts, 0)),

    "Sensible Recovery Efficiency @ 0°C (SRE %)": int(round(sreVal, 0)),
    "Net Moisture Transfer @ 0°C %": int(round(moistVal, 0)),

   "productLink": linkAddress,

        # optional: keep the display text if you want
        "productText": linkText,
            }
        )

    return {"success": True, "results": results, "skippedUnits": skippedUnits, "totalUnits": unitCount}


if __name__ == "__main__":
    path = r"Selection Program For HRV & ERV Compact_Mini_Superior.xlsm"  # update path if needed

    res = get_best_result_per_unit_from_excel(
        excel_path=path,
        airUnit="Cubic Feet Per Minute (CFM)",
        airflow=96,
        spUnit="Pascals (Pa)",
        pressure=29,
        modelType="ERV",
        motorType="PSC",
        sreInput=71,
        moistInput=51,
    )


    from pprint import pprint
    pprint(res)